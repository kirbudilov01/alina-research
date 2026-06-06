from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "output" / "finance" / "AURA_full_financial_model.xlsx"


PURPLE = "9C7AE6"
PURPLE_LIGHT = "EEE8FF"
GREEN_LIGHT = "E2F4DA"
INK = "1E1A2A"
GRID = "8A8A98"
WHITE = "FFFFFF"


def style_sheet(ws):
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin", color=GRID)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Arial", size=10, color=INK)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=PURPLE)
        cell.font = Font(name="Arial", size=13, bold=True, color=WHITE)
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def set_widths(ws, widths):
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def money_format(ws, ranges):
    for rng in ranges:
        for row in ws[rng]:
            for cell in row:
                cell.number_format = '€#,##0;[Red]-€#,##0'


def pct_format(ws, ranges):
    for rng in ranges:
        for row in ws[rng]:
            for cell in row:
                cell.number_format = '0.0%'


def build():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    cover = wb.active
    cover.title = "Read Me"
    cover.append(["AURA Full Financial Model", ""])
    cover.append(["Что это", "Редактируемая модель, чтобы понять порядок инвестиций, выручки, расходов, runway и сценариев роста AURA."])
    cover.append(["Валюта модели", "EUR. Подписка задана в USD и переводится через FX в Inputs."])
    cover.append(["Главное ограничение", "Это hypothesis model, а не прогноз продаж. После прототипа нужно заменить assumptions фактическими данными."])
    cover.append(["Как пользоваться", "Менять желтые/зеленые вводные на вкладке Inputs и смотреть Monthly Model, Quarterly Summary, Annual Scenarios."])
    cover.append(["Что считается", "MAU, платящие, выручка, COGS AI/генераций, маркетинг, команда, burn, cumulative cash need, ARR run-rate."])
    cover.append(["Ключевой вывод", "Phase 1 = €4k-€10k. Полный первый год лучше планировать как €80k-€150k runway до доказанного break-even."])
    set_widths(cover, [28, 110])
    style_sheet(cover)
    cover["A1"].font = Font(name="Arial", size=16, bold=True, color=WHITE)
    for row in range(2, cover.max_row + 1):
        cover[f"A{row}"].fill = PatternFill("solid", fgColor=PURPLE_LIGHT)
        cover[f"A{row}"].font = Font(name="Arial", size=11, bold=True, color=INK)
        cover[f"B{row}"].fill = PatternFill("solid", fgColor=WHITE)

    inputs = wb.create_sheet("Inputs")
    inputs.append(["Input", "Value", "Unit", "Meaning", "Can edit?"])
    rows = [
        ["FX USD to EUR", 0.92, "EUR per USD", "Перевод долларовой подписки в модельную валюту EUR", "yes"],
        ["Base subscription price", 12.99, "USD / month", "Базовая blended цена подписки", "yes"],
        ["App/payment fee", 0.15, "% revenue", "Комиссия store/payment/revenue platform", "yes"],
        ["Token revenue per payer", 1.5, "USD / payer / month", "Средний upside от premium tokens/video", "yes"],
        ["Paid COGS per payer", 3.2, "EUR / payer / month", "AI/image/video COGS для платящего пользователя", "yes"],
        ["Free COGS per free MAU", 0.18, "EUR / free MAU / month", "AI/image COGS free users с capped usage", "yes"],
        ["Support cost per payer", 0.6, "EUR / payer / month", "Support/admin переменная часть", "yes"],
        ["Opening cash", 0, "EUR", "Деньги на старте модели", "yes"],
        ["Year 1 target runway", 120000, "EUR", "Ориентир бюджета первого года", "yes"],
        ["Break-even target month", 12, "month", "Ожидаемый месяц проверки break-even", "yes"],
    ]
    for row in rows:
        inputs.append(row)
    set_widths(inputs, [28, 18, 24, 70, 14])
    style_sheet(inputs)
    for row in range(2, inputs.max_row + 1):
        inputs[f"B{row}"].fill = PatternFill("solid", fgColor=GREEN_LIGHT)
    pct_format(inputs, ["B4:B4"])
    money_format(inputs, ["B9:B10"])

    terms = wb.create_sheet("Terms")
    terms.append(["Term", "Simple explanation", "Why it matters for AURA"])
    term_rows = [
        ["MAU", "Активные пользователи за месяц", "Показывает реальный размер продукта, не просто скачивания"],
        ["Paid conversion", "Какая доля MAU платит", "Главный рычаг выручки"],
        ["Payers", "Количество платящих пользователей", "Основа подписочной выручки"],
        ["COGS", "Переменные расходы на AI, картинки, видео, storage, ошибки генерации", "Если COGS высокий, видео съедает маржу"],
        ["CAC", "Стоимость привлечения клиента", "Нельзя масштабировать, если CAC выше будущей прибыли"],
        ["Burn", "Сколько денег сгорает в месяц", "Показывает, сколько runway нужно"],
        ["Runway / cash need", "Сколько денег нужно, чтобы дожить до следующей проверки", "Помогает планировать раунды/бюджет"],
        ["ARR run-rate", "Месячная выручка × 12", "Удобный ориентир масштаба бизнеса"],
        ["Break-even", "Месяц, когда выручка покрывает расходы", "Показывает шанс стать cashflow-бизнесом"],
        ["Cash-out", "Сценарий продажи/выхода или превращения в денежный бизнес", "Возможен только после доказанных retention, CAC, COGS"],
    ]
    for row in term_rows:
        terms.append(row)
    set_widths(terms, [20, 54, 64])
    style_sheet(terms)

    monthly = wb.create_sheet("Monthly Model")
    headers = [
        "Month", "Phase", "MAU", "Paid conversion", "Payers", "Gross sub revenue",
        "Payment fee", "Net sub revenue", "Token revenue", "Total revenue",
        "Paid COGS", "Free COGS", "Support variable", "Product/dev/team",
        "Infra/tools", "Marketing", "Total spend", "Net cashflow",
        "Cumulative cash need", "ARR run-rate"
    ]
    monthly.append(headers)
    mau = [100, 500, 1000, 3000, 6000, 10000, 15000, 25000, 40000, 60000, 90000, 150000,
           180000, 220000, 270000, 330000, 400000, 500000, 600000, 720000, 850000, 1000000, 1200000, 1400000,
           1600000, 1800000, 2100000, 2400000, 2800000, 3200000, 3600000, 4000000, 4500000, 5000000, 5600000, 6200000]
    conv = [0.0, 0.02, 0.03, 0.027, 0.025, 0.03, 0.03, 0.03, 0.03, 0.03, 0.033, 0.045,
            0.045, 0.046, 0.047, 0.048, 0.049, 0.05, 0.052, 0.054, 0.055, 0.057, 0.058, 0.06,
            0.061, 0.062, 0.063, 0.064, 0.065, 0.066, 0.067, 0.068, 0.069, 0.07, 0.071, 0.072]
    product = [6000, 7000, 12000, 16000, 20000, 24000, 26000, 30000, 33000, 38000, 42000, 50000,
               55000, 60000, 65000, 70000, 75000, 85000, 95000, 105000, 115000, 130000, 145000, 160000,
               175000, 190000, 210000, 230000, 250000, 275000, 300000, 330000, 360000, 390000, 430000, 470000]
    infra = [500, 700, 1000, 1500, 2200, 3000, 3500, 4500, 5500, 7000, 8500, 11000,
             13000, 15000, 18000, 21000, 24000, 28000, 33000, 38000, 44000, 52000, 60000, 70000,
             80000, 90000, 105000, 120000, 140000, 160000, 180000, 205000, 230000, 260000, 295000, 330000]
    marketing = [2000, 3000, 4000, 6000, 9000, 12000, 16000, 22000, 30000, 40000, 50000, 65000,
                 80000, 95000, 110000, 130000, 150000, 180000, 210000, 250000, 300000, 360000, 430000, 500000,
                 580000, 670000, 780000, 900000, 1050000, 1200000, 1380000, 1580000, 1800000, 2050000, 2350000, 2700000]
    phases = ["Validation", "Validation", "MVP build", "MVP build", "Private beta", "Private beta",
              "Public beta", "Public beta", "Public beta", "Launch", "Launch", "Launch"] + ["Scale"] * 24
    for i in range(36):
        r = i + 2
        monthly.append([
            i + 1, phases[i], mau[i], conv[i],
            f"=C{r}*D{r}",
            f"=E{r}*Inputs!$B$3*Inputs!$B$2",
            f"=F{r}*Inputs!$B$4",
            f"=F{r}-G{r}",
            f"=E{r}*Inputs!$B$5*Inputs!$B$2",
            f"=H{r}+I{r}",
            f"=E{r}*Inputs!$B$6",
            f"=(C{r}-E{r})*Inputs!$B$7",
            f"=E{r}*Inputs!$B$8",
            product[i],
            infra[i],
            marketing[i],
            f"=K{r}+L{r}+M{r}+N{r}+O{r}+P{r}",
            f"=J{r}-Q{r}",
            f"=MAX(0,-SUM($R$2:R{r})-Inputs!$B$9)",
            f"=J{r}*12",
        ])
    set_widths(monthly, [9, 16, 12, 14, 12, 16, 13, 15, 14, 15, 12, 12, 14, 16, 12, 12, 14, 14, 17, 14])
    style_sheet(monthly)
    pct_format(monthly, ["D2:D37"])
    money_format(monthly, ["F2:T37"])

    q = wb.create_sheet("Quarterly Summary")
    q.append(["Quarter", "Start month", "End month", "Revenue", "Spend", "Net cashflow", "Ending cash need", "ARR run-rate"])
    for qi in range(12):
        r = qi + 2
        start = qi * 3 + 2
        end = start + 2
        q.append([
            f"Q{qi + 1}", qi * 3 + 1, qi * 3 + 3,
            f"=SUM('Monthly Model'!J{start}:J{end})",
            f"=SUM('Monthly Model'!Q{start}:Q{end})",
            f"=SUM('Monthly Model'!R{start}:R{end})",
            f"='Monthly Model'!S{end}",
            f"='Monthly Model'!T{end}",
        ])
    set_widths(q, [12, 12, 12, 16, 16, 16, 18, 16])
    style_sheet(q)
    money_format(q, ["D2:H13"])

    annual = wb.create_sheet("Annual Scenarios")
    annual.append(["Scenario", "Year 1 ARR", "Year 2 ARR", "Year 3 ARR", "Cash need", "Interpretation"])
    annual_rows = [
        ["Conservative", "$40k-$120k", "$250k-$600k", "$0.8M-$1.5M", "€80k-€150k", "Small cashflow product if CAC stays low and team remains compact"],
        ["Base", "$150k-$400k", "$0.8M-$2M", "$3M-$6M", "€150k-€500k", "Standalone business / seed story if retention and annual plans work"],
        ["Strong", "$0.5M-$1M", "$3M-$6M", "$8M-$15M", "€500k+", "Strategic interest from wellness, astrology, AI companion or creator platforms"],
        ["Stop / pivot", "<$40k", "<$250k", "No retention", "Stop after validation", "Stop if D1 <20%, paid intent <5%, CAC/LTV broken or Canvas causality is unclear"],
    ]
    for row in annual_rows:
        annual.append(row)
    set_widths(annual, [18, 18, 18, 18, 18, 70])
    style_sheet(annual)

    unit = wb.create_sheet("Unit Economics")
    unit.append(["Metric", "Formula / value", "Base result", "Why it matters"])
    unit_rows = [
        ["Net subscription per payer", "=Inputs!B3*Inputs!B2*(1-Inputs!B4)", "=Inputs!B3*Inputs!B2*(1-Inputs!B4)", "Revenue retained after payment/store fee"],
        ["Token revenue per payer", "=Inputs!B5*Inputs!B2", "=Inputs!B5*Inputs!B2", "Optional premium upside"],
        ["Variable COGS per payer", "=Inputs!B6+Inputs!B8", "=Inputs!B6+Inputs!B8", "AI/video/image/support variable cost"],
        ["Contribution per payer", "=B2+B3-B4", "=C2+C3-C4", "What is left before marketing and fixed team costs"],
        ["Free MAU COGS", "=Inputs!B7", "=Inputs!B7", "Free usage must stay capped"],
        ["Break-even payers at €25k fixed/month", "=25000/C5", "=25000/C5", "Rough minimum payer base to cover compact ops"],
    ]
    for row in unit_rows:
        unit.append(row)
    set_widths(unit, [32, 32, 18, 68])
    style_sheet(unit)
    money_format(unit, ["C2:C7"])

    checks = wb.create_sheet("Checks")
    checks.append(["Check", "Formula", "Expected", "Status"])
    checks.append(["Revenue non-negative", "=MIN('Monthly Model'!J2:J37)>=0", "TRUE", '=IF(B2=C2,"OK","CHECK")'])
    checks.append(["Spend non-negative", "=MIN('Monthly Model'!Q2:Q37)>=0", "TRUE", '=IF(B3=C3,"OK","CHECK")'])
    checks.append(["Cash need visible", "=MAX('Monthly Model'!S2:S37)>0", "TRUE", '=IF(B4=C4,"OK","CHECK")'])
    checks.append(["Month count", '=COUNT(\'Monthly Model\'!A2:A37)', 36, '=IF(B5=C5,"OK","CHECK")'])
    set_widths(checks, [28, 36, 18, 16])
    style_sheet(checks)

    dashboard = wb.create_sheet("Dashboard", 1)
    dashboard.append(["AURA Financial Dashboard", "", "", "", ""])
    dashboard.append(["Metric", "Value", "Source", "", ""])
    dashboard.append(["Peak cash need", "='Monthly Model'!S37", "Monthly Model", "", ""])
    dashboard.append(["Year 1 revenue", "='Quarterly Summary'!D5", "Q4 cumulative year view", "", ""])
    dashboard.append(["Month 12 ARR run-rate", "='Monthly Model'!T13", "Monthly Model", "", ""])
    dashboard.append(["Month 12 net cashflow", "='Monthly Model'!R13", "Monthly Model", "", ""])
    dashboard.append(["Break-even note", "Base-case break-even likely around Q4 / Q1 next year, not in first 3 months.", "", "", ""])
    set_widths(dashboard, [28, 22, 36, 18, 18])
    style_sheet(dashboard)
    money_format(dashboard, ["B3:B6"])

    chart = LineChart()
    chart.title = "Revenue vs Spend"
    chart.y_axis.title = "EUR"
    chart.x_axis.title = "Month"
    data = Reference(monthly, min_col=10, max_col=17, min_row=1, max_row=37)
    cats = Reference(monthly, min_col=1, min_row=2, max_row=37)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 18
    dashboard.add_chart(chart, "A10")

    order = ["Read Me", "Dashboard", "Inputs", "Terms", "Monthly Model", "Quarterly Summary", "Annual Scenarios", "Unit Economics", "Checks"]
    wb._sheets = [wb[name] for name in order]

    for ws in wb.worksheets:
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 24
        ws.row_dimensions[1].height = 32

    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    build()
